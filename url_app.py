#!/usr/bin/env python3
'''This application is  a tkinter app for saving useful online material'''

#------Importing the useful modules------#
'''
   tkinter is an intermediate tool for using tcl,imported tkinter as tk
   openpyxl is an module for working with spread sheet with xlsx/xlsm/xltx/xltm files, imported openpyxl as xl
   pyperclip, a module for inteacting with the clipboard
   web_open is a file containg the piece of code for opening the webbrowser
'''

import tkinter as tk
import openpyxl as xl
from pyperclip import copy,paste
from openpyxl.styles import colors,Font,Alignment,NamedStyle
from web_open import search

#----------------------------------------#

flag = 0 # A variable for checking whether the user has entered the description
def add_link():
    '''The purpose of this method is to open a new toplevel widget add provide an entry field for the usr to enter the description of the link
       he/she intended to save.
    '''

# StringVar is a special variable for holdong the string, it can be assinged to entry widget and the information enterd in the entry widget can b acessed using va1_.get()
    var_1 = tk.StringVar()

#-----Creating a new toplevel widget for user to enter the description-----#
    add = tk.Toplevel()
    add.geometry("800x190")
    add.title("Add a new link")
#--------------------------------------------------------------------------#


    prompt=tk.Label(add,text = "Enter the descriptipon of the link below, make sure you copied the link to the clip board")
    discription = tk.Entry(add,width = 50,textvariable=var_1)
    prompt1 = tk.Label(add,text = "Click to save the link")

#-----Placing the label,entry widgets on the toplevel widget-----#
    discription.focus_set()
    discription.grid(row = 1,column = 0)
    prompt.grid(row = 0,column = 0)
    prompt1.grid()
#----------------------------------------------------------------#

    def check():
        '''
        The purpose of this function is to check whether the user has entered the description, if not it shows a warning to the user
        If the user had entered the description, the toplevel widget is destroyed and the  adding function is called
        '''
        global flag
        if not var_1.get() and not flag:
            warn = tk.Label(add,text = "*Please enter the description",fg = "red")
            warn.grid()
            flag = 1
        elif len(var_1.get())>0:
            c = var_1.get()
            add.destroy()
            adding(c)

    add_button =tk.Button(add,text = "Save",fg = "white",bg ="green",relief = "solid",command = check)
    add_button.grid()

def adding(c):
    '''
       The purpose of this method is to save the link in a spread sheet, first the spread sheet is ceated if it doesn't exist using the openpyxl
       module and the link and description of the link are added to the spead sheet, after adding the spread sheet columns width is adjusted
       based on the lenght of the largest description.
    '''

#----- A toplevel widget for letting the user know that the link has been saved succesfully -----#
    confirm = tk.Toplevel()
    confirm.geometry("400x100")
#------------------------------------------------------------------------------------------------#

    if not len(paste()):
        # to warn the user if he didnt copy any link
        warn = tk.Label(confirm,text = "*Please copy the link",fg = "red")
        warn.grid()
    else:
        pos =tk.Label(confirm,text = "Your link has been added",fg= "green")
        pos.grid()
        close = tk.Button(confirm,text ="OK",bg = "blue",fg = "white",command =  confirm.destroy)
        close.grid()
    try:
        #-----To open an existing spread sheet if exist-----#
        wb = xl.load_workbook("Saved_links.xlsx")
        sheet = wb.get_sheet_by_name("saved_links")

    except:
        #------If the spread sheet doesn't exist, a new workbook is created-------#
        wb = xl.Workbook()
        sheet = wb.create_sheet()
        sheet.title = "saved_links"
        a1 = sheet['A1']
        b1 = sheet['B1']
        #To change the font color and the style of the head of the columns
        ft = Font(color = colors.BLACK, bold = True)
        # Aligning the heading of the column to centre
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical = 'center')
        sheet['B1'].alignment = Alignment(horizontal = 'center',vertical = 'center')
        a1.font = ft
        b1.font = ft
        a1.value = "Description of the link"
        b1.value = "Link"
    #saving the changes made to the spread sheet
    wb.save("Saved_links.xlsx")


#------------------- Finding the next immediate empty row ------------------#
    empty_row_number = 0
    for cell in sheet['A']:
        if not cell.value :
            empty_row_number = cell.row
            break
    else:
        #----- If no empty row exist, a new row is created -----#
         empty_row_number = ((cell.row) + 1)
#--------------------------------------------------------------------------#

#------ Storing the link and description in the spread sheet -----#
    sheet[f'A{empty_row_number}'].value = c

    

    sheet[f'B{empty_row_number}'].value = paste()
    copy("")# emptying the clipboard
    #saving the spread sheet
    wb.save("Saved_links.xlsx")
#---------------------------------------------------------------------------#


#---------- TO FIND OUT THE MAXIMUM LENGTH OF THE DESCRIPTION AND LINK, ACCORDINGLY ADJUSTING THE WIDTH OF THE RESPECTIVE COLUMN---------#
    max_width = 0

    # finding out the maximum length of the description
    for cell in sheet['A']:
        if len(cell.value) > max_width:
            max_width = len(cell.value)
    # changing the description column width
    sheet.column_dimensions["A"].width = max_width + 4
    # saving the changes made
    wb.save("Saved_links.xlsx")
    # similarly finding out the max_width of the link column
    max_width = 0
    for cell in sheet['B']:
        if len(cell.value) > max_width:
            max_width = len(cell.value)
    sheet.column_dimensions['B'].width = max_width + 4

    # saving the changes made
    wb.save("Saved_links.xlsx")

#----------------------------------------------------------The changes made to sheet dimensions end here---------------------------------#


def show():
    show_ = tk.Toplevel()


        #This function displays the link so far saved, user can either delete the link or view the link in webbrowser



#------- Setting the toplevel widget with scrollbar and a listbox --------#
    # creating a toplevel widget

    show_.title("Saved information")
    show_.geometry("550x350")
    can_ = tk.Frame(show_,width =500,height = 500)
    label = tk.Label(can_,text = "These  are  the  description of the links you saved :)",font = 6)
    #settng the scrollbar, inorder to use a scrollbar we need to use a frame
    scrollbar = tk.Scrollbar(can_,orient = tk.VERTICAL)#assinging the orietation vertically
    listbox = tk.Listbox(can_,selectmode = tk.EXTENDED,yscrollcommand = scrollbar.set)#binding the scrollbar to the frame along the y-axis
#------------------------------------------------------------------------#

    label.pack()
    scrollbar.config(command = listbox.yview)
    scrollbar.pack(side = tk.RIGHT,fill = tk.Y)
    wb = xl.load_workbook("Saved_links.xlsx")
    sheet = wb.get_sheet_by_name("saved_links")
    for row in range(2,sheet.max_row + 1):
        #filling up the list box with the description data from the spread sheet
         listbox.insert(tk.ACTIVE,(sheet[f'A{row}'].value))


    def delete():
        '''
           To provide the user delete option to permanently delete the selected option
        '''
        items = list(map(int,listbox.curselection()))#this gives the index position starting from 0 of the list elements selected, it is a string, so we are converetng it into  an integer

        if(len(items))==0:
            '''
               If the user hasn't selected the any option and clicked the delete/open button the tolevel gets destroyed
            '''
            show_.destroy()
            exit

        for i in items:
            try:
                # deleting the selected description and link from the spread sheet, based on the index position
                sheet.delete_rows(i+2)
                wb.save("Saved_links.xlsx")
            except:
                exit("Didn't work")
            listbox.delete(i)



    def open():
        '''
           opening the link in a browser
        '''
        items = list(map(int,listbox.curselection()))
        flag = 1
        if(len(items))==0:

            show_.destroy()
        else:
            for i in items:
                col = int(i) + 2
                search(str(sheet[f'B{col}'].value))

    delete1 = tk.Button(show_,text = "Delete",fg = "green",bg = "white",command = lambda:delete())
    open1 = tk.Button(show_,text = "Open",fg ="green",bg="white",command = lambda:open())


    delete1.pack(side =tk.BOTTOM)
    open1.pack(side = tk.BOTTOM)
    listbox.pack(side = tk.LEFT,fill=tk.BOTH,expand = 1)
    can_.pack()

#---------------------------------------------------------------------------------#

'''
    Creating a tkinter object and assinging it to a variable named 'window' .
    Then setting the title method of window to 'Url Saver', it appears on the top whenever the app is launched.
    Setting the geometry of the window to required size
    Setting the app icon
    Associating a method 'add_link' to a button so that when user clicks the button a new toplevel appears for saving the URL
    Associating a method 'show' to a button so that user can see all the saved url and can either open or delete the url
'''

window =tk.Tk()
window.title("Url Saver")
window.geometry("400x100")

#------ Setting up the app icon object -----#
img = tk.Image("photo",file = "app_icon.png")
window.tk.call('wm','iconphoto',window._w,img)
#-------------------------------------------#

w = tk.Label(text = "Hello, Welcome to URL saver app!!")

#----- Placing the welcome label and buttons on the main/root window ------#
w.grid()
button = tk.Button(window,text = "Add a new link",fg="white",bg="green",command = add_link)
button1 = tk.Button(window,text= "Search for the link",fg = "white",bg="green",command = show)
button.grid()
button1.grid(row = 2)
#-------------------------------------------------------------------------#


#----- This puts the main window on screen -----#
window.mainloop()
#------------------------------The SCRIPT ENDS HERE------------------------------#
